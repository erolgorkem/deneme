import os
import sqlite3
from flask import Flask, request, render_template_string, redirect, url_for, session, g
import openpyxl
from tempfile import NamedTemporaryFile

app = Flask(__name__)
app.secret_key = 'gorkem-bey-ozel-sifre'

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, "data.db")

SIPARIS_KOLONLAR = [
    "Sipariş Statüsü", "Sipariş Tarihi", "Teslim Tarihi", "Sipariş Numarası", "Barkod",
    "Stok Kodu", "Adet", "Alıcı", "Paket No", "Kargo Firması", "Kargo Kodu", "Ürün Adı",
    "Birim Fiyatı", "Satış Tutarı", "İndirim Tutarı", "Faturalanacak Tutar", "Komisyon Oranı",
    "Teslimat Adresi", "İl", "İlçe"
]
MALIYET_KOLONLAR = [
    "Barkod", "Model Kodu", "Stok Kodu", "Kategori", "Ürün Adı",
    "Trendyol Satış Fiyatı", "Ürün Maliyeti (KDV Dahil)"
]
KULLANICI_ADI = "admin"
SIFRE = "bilmiyorum"

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
    return db

def init_db():
    db = get_db()
    c = db.cursor()
    siparis_fields = ', '.join([f'"{k.replace(" ", "_")}" TEXT' for k in SIPARIS_KOLONLAR])
    maliyet_fields = ', '.join([f'"{k.replace(" ", "_")}" TEXT' for k in MALIYET_KOLONLAR])
    c.execute(f'CREATE TABLE IF NOT EXISTS siparisler (id INTEGER PRIMARY KEY AUTOINCREMENT, {siparis_fields})')
    c.execute(f'CREATE TABLE IF NOT EXISTS maliyetler (id INTEGER PRIMARY KEY AUTOINCREMENT, {maliyet_fields})')
    db.commit()

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def kolon_eslestir(headers, aranan_kolonlar):
    # Excel başlıklarını normalize et ve eşle
    temiz = lambda s: ''.join(str(s).lower().replace(' ', '').replace('\n','').replace('-','').replace('(','').replace(')','').replace('_',''))
    header_map = {temiz(col): i for i, col in enumerate(headers)}
    sonuc = []
    for aranan in aranan_kolonlar:
        anahtar = temiz(aranan)
        sonuc.append(header_map.get(anahtar, None))
    return sonuc

@app.route('/login', methods=['GET', 'POST'])
def login():
    hata = None
    if request.method == 'POST':
        kullanici = request.form.get('kullanici')
        sifre = request.form.get('sifre')
        if kullanici == KULLANICI_ADI and sifre == SIFRE:
            session['giris'] = True
            return redirect(url_for('index'))
        else:
            hata = "Kullanıcı adı veya şifre yanlış!"
    return render_template_string("""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Giriş Yap</title>
            <meta charset="utf-8">
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
            <style>
                body { background-color: #f4f6f8; }
                .container { margin-top: 120px; }
            </style>
        </head>
        <body>
            <div class="container col-md-4 mx-auto text-center">
                <h2 class="text-primary mb-4">Giriş Yap</h2>
                <form method="post" class="border rounded-4 p-5 shadow bg-white">
                    <input type="text" name="kullanici" placeholder="Kullanıcı Adı" class="form-control mb-3" required autofocus>
                    <input type="password" name="sifre" placeholder="Şifre" class="form-control mb-4" required>
                    {% if hata %}
                        <div class="alert alert-danger mb-3">{{ hata }}</div>
                    {% endif %}
                    <button type="submit" class="btn btn-success btn-lg rounded-pill px-5">Giriş</button>
                </form>
            </div>
        </body>
        </html>
    """, hata=hata)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def index():
    if not session.get('giris'):
        return redirect(url_for('login'))
    return redirect(url_for('siparis'))

def excel_satir_satir_database(tabloadi, kolonlar, file_storage):
    # Yüklenen dosyayı RAM'e almadan geçici dosyaya kaydet
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file_storage.save(tmp.name)
        tmp_path = tmp.name
    db = get_db()
    c = db.cursor()
    # Tabloyu temizle
    c.execute(f"DELETE FROM {tabloadi}")
    wb = openpyxl.load_workbook(tmp_path, read_only=True)
    ws = wb.active
    # Başlıklar ve kolon eşleşmeleri
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    kolon_idx = kolon_eslestir(headers, kolonlar)
    # INSERT cümlesi
    sql_cols = ', '.join([f'"{k.replace(" ","_")}"' for k in kolonlar])
    sql_query = f'INSERT INTO {tabloadi} ({sql_cols}) VALUES ({",".join(["?"]*len(kolonlar))})'
    sayac = 0
    for row in ws.iter_rows(min_row=2):
        values = []
        for idx in kolon_idx:
            if idx is not None and idx < len(row):
                val = row[idx].value
            else:
                val = ""
            values.append(str(val) if val is not None else "")
        c.execute(sql_query, values)
        sayac += 1
        if sayac % 100 == 0:
            db.commit()
    db.commit()
    wb.close()
    os.unlink(tmp_path)
    return sayac

@app.route('/siparis', methods=['GET', 'POST'])
def siparis():
    if not session.get('giris'):
        return redirect(url_for('login'))
    hata = None
    if request.method == 'POST':
        file = request.files['file']
        if file and file.filename.endswith(('.xls', '.xlsx')):
            try:
                eklenen = excel_satir_satir_database('siparisler', SIPARIS_KOLONLAR, file)
            except Exception as e:
                hata = f"Excel yüklemede hata: {str(e)}"
    db = get_db()
    sql_cols = ', '.join([f'"{k.replace(" ","_")}"' for k in SIPARIS_KOLONLAR])
    rows = db.execute(f"SELECT {sql_cols} FROM siparisler").fetchall()
    tablo_df = rows if rows else None
    return render_sablon(
        aktif_tab="siparis",
        tablo_df=tablo_df,
        kolonlar=SIPARIS_KOLONLAR,
        yukleme_hatasi=hata
    )

@app.route('/maliyet', methods=['GET', 'POST'])
def maliyet():
    if not session.get('giris'):
        return redirect(url_for('login'))
    hata = None
    if request.method == 'POST':
        file = request.files['file']
        if file and file.filename.endswith(('.xls', '.xlsx')):
            try:
                eklenen = excel_satir_satir_database('maliyetler', MALIYET_KOLONLAR, file)
            except Exception as e:
                hata = f"Excel yüklemede hata: {str(e)}"
    db = get_db()
    sql_cols = ', '.join([f'"{k.replace(" ","_")}"' for k in MALIYET_KOLONLAR])
    rows = db.execute(f"SELECT {sql_cols} FROM maliyetler").fetchall()
    tablo_df = rows if rows else None
    return render_sablon(
        aktif_tab="maliyet",
        tablo_df=tablo_df,
        kolonlar=MALIYET_KOLONLAR,
        yukleme_hatasi=hata
    )

def render_sablon(aktif_tab, tablo_df, kolonlar, yukleme_hatasi=None):
    return render_template_string("""
    <!DOCTYPE html>
    <html>
    <head>
      <title>Excel Paneli</title>
      <meta charset="utf-8">
      <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
      <style>
        body { background-color: #f4f6f8; }
        .container { margin-top: 30px; max-width: 99vw; }
        .compact-table { font-size: 0.90rem; }
        .compact-table th, .compact-table td {
          padding-top: 3px !important; padding-bottom: 3px !important;
          padding-left: 5px !important; padding-right: 5px !important;
          white-space: nowrap;
        }
        .table { border-radius: 12px; overflow: hidden; }
        th { background: #007bff; color: #fff; cursor: pointer;}
        td, th { vertical-align: middle !important; }
        .table-responsive { max-height: 65vh; overflow-x: auto; }
        .arama-kutusu { margin-bottom: 12px; max-width: 320px;}
        .tabnav { margin-bottom: 14px; }
        .yukle-kart {max-width:450px; margin:auto; margin-bottom:15px;}
        .gizli {display:none;}
      </style>
      <script>
        function filtreleTablo() {
          var input = document.getElementById("aramaInput");
          var filtre = input.value.toLowerCase();
          var tablo = document.getElementById("veriTablosu");
          var tr = tablo ? tablo.getElementsByTagName("tr") : [];
          for (var i = 1; i < tr.length; i++) {
            var tds = tr[i].getElementsByTagName("td");
            var satirGoster = false;
            for (var j = 0; j < tds.length; j++) {
              if (tds[j].textContent.toLowerCase().indexOf(filtre) > -1) {
                satirGoster = true; break;
              }
            }
            tr[i].style.display = satirGoster ? "" : "none";
          }
        }
      </script>
    </head>
    <body>
      <div class="container-fluid">
        <div style="position:absolute;top:18px;right:28px;">
            <a href="/logout" class="btn btn-outline-danger">Çıkış Yap</a>
        </div>
        <ul class="nav nav-tabs tabnav">
          <li class="nav-item">
            <a class="nav-link {% if aktif_tab == 'siparis' %}active{% endif %}" href="/siparis">Sipariş Exceli</a>
          </li>
          <li class="nav-item">
            <a class="nav-link {% if aktif_tab == 'maliyet' %}active{% endif %}" href="/maliyet">Maliyet Exceli</a>
          </li>
        </ul>
        <div class="card yukle-kart shadow-sm border-0">
          <div class="card-body p-3">
            <h5 class="card-title text-primary mb-3" style="font-size:1.1rem;">
              {% if aktif_tab == 'siparis' %}Sipariş Exceli Yükle{% else %}Maliyet Exceli Yükle{% endif %}
            </h5>
            <form method="post" enctype="multipart/form-data">
              <input type="file" name="file" accept=".xls,.xlsx" class="form-control form-control-sm mb-2" required>
              <button type="submit" class="btn btn-primary btn-sm rounded-pill px-4">Yükle</button>
            </form>
            {% if yukleme_hatasi %}
              <div class="alert alert-danger mt-2 mb-0 p-2" style="font-size:0.96em;">{{ yukleme_hatasi }}</div>
            {% endif %}
          </div>
        </div>
        <div class="mt-3 mb-2 {% if tablo_df is none %}gizli{% endif %}">
          <input type="text" id="aramaInput" class="form-control arama-kutusu" placeholder="Tabloda ara..." onkeyup="filtreleTablo()">
        </div>
        <div class="table-responsive shadow rounded-4" style="background:white;">
          {% if tablo_df is not none %}
            <table class="table table-hover table-bordered align-middle text-center compact-table" id="veriTablosu">
              <thead>
                <tr>
                  {% for kol in kolonlar %}
                    <th>{{kol}}</th>
                  {% endfor %}
                </tr>
              </thead>
              <tbody>
                {% for row in tablo_df %}
                  <tr>
                    {% for value in row %}
                      <td>{{ value }}</td>
                    {% endfor %}
                  </tr>
                {% endfor %}
              </tbody>
            </table>
          {% endif %}
        </div>
      </div>
    </body>
    </html>
    """, tablo_df=tablo_df, kolonlar=kolonlar, aktif_tab=aktif_tab, yukleme_hatasi=yukleme_hatasi)

# Tabloları oluştur
with app.app_context():
    init_db()

if __name__ == '__main__':
    app.run(debug=True)
